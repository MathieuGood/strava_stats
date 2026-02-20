import io
import os
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from .config import RATE_PER_KM


class CommuteReport:
    def __init__(self, commute_activities, year, month):
        self.activities = commute_activities
        self.year = year
        self.month = month

    def _group_by_date(self):
        """Group activities by date, summing distances and collecting trips."""
        days = defaultdict(lambda: {"distance_km": 0.0, "trips": []})
        for a in self.activities:
            d = a["date"]
            days[d]["distance_km"] += a["distance_km"]
            days[d]["trips"].append(a)
        return dict(sorted(days.items()))

    def _build_workbook(self) -> Workbook:
        wb = Workbook()
        ws = wb.active
        ws.title = "Indemnité km vélo"

        arial = Font(name="Arial", size=11)
        header_font = Font(name="Arial", size=11, bold=True)
        total_font = Font(name="Arial", size=11, bold=True)
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Column headers
        headers = [
            "Date", "Jour", "Trajet Aller", "Trajet Retour",
            "Motif", "Distance (km)", "Indemnité/km", "Indemnité (€)"
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        days = self._group_by_date()
        row = 2
        for date, info in days.items():
            trips = info["trips"]
            # Aller = first trip of the day, Retour = last trip (if different)
            aller = f"{trips[0]['departure']} → {trips[0]['arrival']}"
            retour = ""
            if len(trips) > 1:
                retour = f"{trips[-1]['departure']} → {trips[-1]['arrival']}"

            def c(col):
                cell = ws.cell(row=row, column=col)
                cell.font = arial
                cell.border = thin_border
                return cell

            c(1).value = date
            c(1).number_format = "DD/MM/YYYY"
            c(2).value = f'=TEXT(A{row},"jjjj")'
            c(3).value = aller
            c(4).value = retour
            c(5).value = "Trajet domicile-travail"
            c(6).value = round(info["distance_km"], 2)
            c(6).number_format = "0.00"
            c(7).value = RATE_PER_KM
            c(7).number_format = "0.00"
            c(8).value = f"=F{row}*G{row}"
            c(8).number_format = "0.00"

            row += 1

        # TOTAL row
        total_row = row
        ws.cell(row=total_row, column=5, value="TOTAL").font = total_font
        ws.cell(row=total_row, column=5).border = thin_border
        ws.cell(row=total_row, column=5).alignment = Alignment(horizontal="right")
        ws.cell(row=total_row, column=6).value = f"=SUM(F2:F{total_row - 1})"
        ws.cell(row=total_row, column=6).number_format = '0.00" km"'
        ws.cell(row=total_row, column=6).font = total_font
        ws.cell(row=total_row, column=6).border = thin_border
        ws.cell(row=total_row, column=8).value = f"=SUM(H2:H{total_row - 1})"
        ws.cell(row=total_row, column=8).number_format = '0.00" €"'
        ws.cell(row=total_row, column=8).font = total_font
        ws.cell(row=total_row, column=8).border = thin_border

        # Column widths
        col_widths = [14, 12, 28, 28, 24, 14, 14, 14]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = w

        return wb

    def generate(self, output_dir="reports"):
        os.makedirs(output_dir, exist_ok=True)
        filename = f"Indemnité_KM_mobilite_velo_MB_{self.year}_{self.month:02d}.xlsx"
        filepath = os.path.join(output_dir, filename)
        self._build_workbook().save(filepath)
        return filepath

    def generate_to_bytes(self) -> bytes:
        """Generate the workbook and return raw bytes (for HTTP streaming)."""
        buffer = io.BytesIO()
        self._build_workbook().save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
